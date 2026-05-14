#!/usr/bin/env python3
"""
Lee Ventas Mambula *.xlsx (pestaña VENTAS; Promocionales opcional) y escribe SQL
DELETE public.sales + INSERT + opcional UPDATE promocional_rows.

Mapeo de columnas (encabezados, sin distinguir mayúsculas / espacios laterales):
  Consumidor → buyer
  Unidades → quantity
  Precio → unit_price_ars
  Total → referencia (cant × precio); si diff grande solo advertimos en cabecera SQL
  Vendedor → seller
  Pagado → paid_ars (valor en Excel)
  Entregado → delivered (SI → SI)
  Facturado → invoice_status (SI → facturado; vacío → pendiente)
  Notas → billing_notes

payment_method / payment_status se infieren como antes (notas + monto vs total; parcial si hay pago sin cerrar).

Uso:
  python3 scripts/generate_sync_from_excel.py "/ruta/archivo.xlsx" [ruta_salida.sql] [sold_at YYYY-MM-DD]
"""

from __future__ import annotations

import json
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = ROOT / "supabase/migrations/0023_resync_sales_from_excel.sql"

SOC_PROMO = frozenset({"Delfi", "Mechi", "Susan"})


def esc_sql(s: str | None) -> str:
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def norm_header(h) -> str:
    if h is None:
        return ""
    return str(h).strip().lower()


def num_cell(v) -> int:
    if v is None:
        return 0
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int,)):
        return int(v)
    if isinstance(v, float):
        return int(round(v))
    if isinstance(v, Decimal):
        return int(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s:
        return 0
    s = s.replace(".", "").replace(",", ".") if "," in s and "." not in s else s.replace(",", "")
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


def infer_payment_method(notes: str | None, billing_parts: list[str]) -> str:
    blob = " ".join([notes or ""] + billing_parts).lower()
    if "efectivo" in blob:
        return "efectivo"
    if "mercado" in blob or "transfer" in blob or "mambula" in blob or "cuenta" in blob:
        return "transferencia"
    return "otro"


def invoice_status_from_fact(fact) -> str:
    if fact is None or str(fact).strip() == "":
        return "pendiente"
    fs = str(fact).strip().upper()
    if "SIN FACTURA" in fs:
        return "pendiente"
    if fs.startswith("SI"):
        return "facturado"
    return "pendiente"


def delivered_cell(ent) -> str | None:
    if ent is None or str(ent).strip() == "":
        return None
    s = str(ent).strip()
    if s.upper() == "SI":
        return "SI"
    return s


def build_column_index(header_row: tuple) -> dict[str, int]:
    idx: dict[str, int] = {}
    aliases = {
        "consumidor": "consumidor",
        "unidades": "unidades",
        "precio": "precio",
        "total": "total",
        "vendedor": "vendedor",
        "pagado": "pagado",
        "entregado": "entregado",
        "facturado": "facturado",
        "notas": "notas",
    }
    for i, cell in enumerate(header_row):
        key = norm_header(cell)
        if key in aliases:
            idx[aliases[key]] = i
    required = ("consumidor", "unidades", "precio", "vendedor", "pagado", "entregado", "facturado")
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f"Faltan columnas en VENTAS: {missing}. Fila 1: {header_row!r}")
    return idx


def parse_ventas(ws) -> list[tuple]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    idx = build_column_index(rows[0])
    i_cons = idx["consumidor"]
    i_qty = idx["unidades"]
    i_precio = idx["precio"]
    i_total = idx.get("total")
    i_seller = idx["vendedor"]
    i_pagado = idx["pagado"]
    i_ent = idx["entregado"]
    i_fact = idx["facturado"]
    i_notas = idx.get("notas")

    out: list[tuple] = []
    for r in rows[1:]:
        if not r or len(r) <= max(idx.values()):
            continue
        buyer_cell = r[i_cons]
        if buyer_cell is None or str(buyer_cell).strip() == "":
            continue
        buyer = str(buyer_cell).strip()
        if buyer.upper() == "TOTAL":
            continue

        qty = r[i_qty]
        unit_p = r[i_precio]
        total_x = r[i_total] if i_total is not None else None
        seller = str(r[i_seller]).strip() if r[i_seller] else ""
        pagado = r[i_pagado]
        entregado = r[i_ent]
        facturado = r[i_fact]
        notas = ""
        if i_notas is not None and len(r) > i_notas and r[i_notas]:
            notas = str(r[i_notas]).strip()

        q = num_cell(qty) if qty is not None else None
        if q is None or q <= 0:
            continue

        if not seller:
            seller = "Susan"

        up = num_cell(unit_p) if unit_p is not None else None
        if up is None or up <= 0:
            up = 15000

        total_calc = q * up
        tx = num_cell(total_x) if total_x is not None else 0
        if tx > 0 and abs(tx - total_calc) > max(total_calc * 0.15, 1):
            pass

        paid_raw = max(0, num_cell(pagado))

        inv = invoice_status_from_fact(facturado)

        billing_parts: list[str] = []
        if notas:
            billing_parts.append(notas)
        if facturado and str(facturado).strip() and str(facturado).strip().upper() not in ("SI", "TRUE"):
            billing_parts.append(str(facturado).strip())
        elif facturado and "," in str(facturado):
            billing_parts.append(str(facturado).strip())

        billing_note = " · ".join(billing_parts) if billing_parts else None

        pay_method = infer_payment_method(notas, billing_parts)

        delivered = delivered_cell(entregado)

        if total_calc > 0 and paid_raw >= total_calc:
            pay_stat = "cobrado"
        elif paid_raw > 0:
            pay_stat = "parcial"
        else:
            pay_stat = "pendiente"

        paid_adj = paid_raw

        out.append(
            (
                buyer,
                seller if seller else None,
                q,
                up,
                pay_method,
                pay_stat,
                paid_adj,
                delivered,
                billing_note,
                inv,
            )
        )
    return out


def parse_promos(ws) -> dict[str, list[dict]]:
    rows = list(ws.iter_rows(values_only=True))
    groups = {"equipo": [], "colaboracion": [], "influencers": [], "colegio": []}
    group = "equipo"
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        a = str(r[0]).strip()
        if not a:
            continue
        al = a.lower()
        if al == "total":
            continue
        if "colab" in al:
            group = "colaboracion"
            continue
        if al == "influencers":
            group = "influencers"
            continue
        if "colegio" in al:
            group = "colegio"
            continue
        if al in ("equipo",):
            continue

        unidades = num_cell(r[1]) if len(r) > 1 else 1
        por = r[2] if len(r) > 2 else None
        por_s = str(por).strip() if por is not None else ""
        if por_s in SOC_PROMO:
            entregado = True
            entregado_por = por_s
        else:
            entregado = False
            entregado_por = None

        groups[group].append(
            {
                "nombre": a,
                "unidades": max(1, unidades),
                "entregado": entregado,
                "entregadoPor": entregado_por,
            }
        )
    return groups


def main():
    xlsx = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else None
    out_path = Path(sys.argv[2]).expanduser() if len(sys.argv) > 2 else DEFAULT_OUT
    sold_at = sys.argv[3] if len(sys.argv) > 3 else "2026-05-06"

    if not xlsx or not xlsx.is_file():
        print(
            "Uso: python3 scripts/generate_sync_from_excel.py /ruta/al/archivo.xlsx [salida.sql] [YYYY-MM-DD]",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if "VENTAS" not in wb.sheetnames:
        wb.close()
        print("El libro debe tener pestaña VENTAS.", file=sys.stderr)
        sys.exit(1)

    ventas = parse_ventas(wb["VENTAS"])
    had_promos_sheet = "Promocionales" in wb.sheetnames
    if had_promos_sheet:
        promos = parse_promos(wb["Promocionales"])
    else:
        promos = None

    wb.close()

    lines = [
        "-- Reemplazo total de public.sales desde Excel (Ventas Mambula).",
        "-- Generado por scripts/generate_sync_from_excel.py",
        f"-- Fuente: {xlsx.name}",
        f"-- sold_at unificado: {sold_at}",
        "",
        "delete from public.sales;",
        "",
        "insert into public.sales (",
        "  sold_at, buyer, seller, quantity, unit_price_ars,",
        "  payment_method, payment_status, paid_ars, delivered, billing_notes, invoice_status,",
        "  sheet_position",
        ") values",
    ]

    value_lines = []
    for pos, row in enumerate(ventas, start=1):
        buyer, seller, q, up, pm, ps, paid, deliv, bill, inv = row
        value_lines.append(
            "  ("
            + f"{esc_sql(sold_at)}::date, "
            + f"{esc_sql(buyer)}, "
            + (esc_sql(seller) if seller else "NULL")
            + ", "
            + f"{q}, {up}, "
            + f"{esc_sql(pm)}, {esc_sql(ps)}, {paid}, "
            + (esc_sql(deliv) if deliv else "NULL")
            + ", "
            + (esc_sql(bill) if bill else "NULL")
            + ", "
            + f"{esc_sql(inv)}, "
            + f"{pos}"
            + ")"
        )

    lines.append(",\n".join(value_lines) + ";")
    lines.append("")

    if had_promos_sheet and promos is not None:
        promo_json = json.dumps(promos, ensure_ascii=False, separators=(",", ":"))
        lines.append(
            "update public.project_settings ps\nset promocional_rows = "
            + "'" + promo_json.replace("'", "''") + "'::jsonb\nwhere ps.id = (select id from public.project_settings limit 1);"
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    extra = f"; promo actualizado: sí ({ {k: len(v) for k, v in promos.items()} })" if had_promos_sheet and promos else "; pestaña Promocionales ausente — promocional_rows no modificado"
    print(f"Escrito {out_path} ({len(ventas)} ventas{extra}).")


if __name__ == "__main__":
    main()
